import openpyxl
# ライブラリopenpyxlを導入
file="sample.xlsx" # ファイル名を何度も入力する必要がないように変数に代入

def excel_update(): # defを使って関数化　こちらは既存のエクセルファイルの読み込み
    wb=openpyxl.load_workbook(file) # openpyxlを使って、変数"file"のファイル名を持つエクセルファイルを読み込む
    ws=wb['Sheet1'] # sample.xlsxのエクセルファイルの中のSheet1というシートを指定し、変数wsに代入
    a=ws['A1'].value # Sheet1のA1セル内の値を、変数aに代入
    print(a) # print関数で変数aにA1セルの値が代入されていることを確認
    return a # 取得したセルデータを次の処理に渡すため、値を返す

def excel_renew(date,time): # こちらはスクレイピング結果を受け取って、エクセルデータを更新する処理を関数化したもの
    wb=openpyxl.load_workbook(file)
    ws=wb['Sheet1']
    ws['B1']=date # wsにエクセルシートのデータが入っており、[]でセルの場所を指定すると、そのセルの値を編集できる
    ws['C1']=time
    wb.save('sample_updated.xlsx') # ここでエクセルファイルを保存。念のためファイル名を分けましたが、同じでもOK
    print('エクセル更新完了しました！') # 最後に処理完了を表示

# if __name__ == '__main__': #作成した関数を実行するための処理
#     excel_update()